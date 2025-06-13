from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import re
import io
import numpy as np
from common import parse_rccl_tests_output

def read_file_as_string(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    return content
    data = []
    regexstr1 = r"\s*(-?\d+)\s+(-?\d+)\s+(\S+)\s+(\S+)\s+(-?\d+)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+((-?\d+)|N/A)"
    regexstr2 = r"\s*(-?\d+)\s+(-?\d+)\s+(\S+)\s+(-?\d+)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+((-?\d+)|N/A)"
    pattern1 = re.compile(regexstr1)
    pattern2 = re.compile(regexstr2)
    for line in io.StringIO(rccl_tests_log_str):
        # Skip lines that start with '##' or don't match the regex pattern
        if line.startswith('##') or not (pattern1.match(line) or pattern2.match(line)):
            continue

        # Split line into columns (handle variable spacing)
        columns = line.split()
            
        # Ensure the line has the expected number of fields
        if len(columns) >= 12:
            i = 1 if pattern1.match(line) else 0
            entry = {
                "size": np.int64(columns[0]),
                "elements": np.int64(columns[1]),
                "type": columns[2],
                "redop": columns[2+i] if i else 'none',
                "root": np.int64(columns[3+i]),
                "op_time(us)": float(columns[4+i]),
                "op_algbw(GB/s)": float(columns[5+i]),
                "op_busbw(GB/s)": float(columns[6+i]),
                "op_wrong": np.int64(columns[7+i]) if columns[7+i].isdigit() or (columns[7+i][1:].isdigit() if columns[7+i].startswith('-') else False) else 0,
                "ip_time(us)": float(columns[8+i]),
                "ip_algbw(GB/s)": float(columns[9+i]),
                "ip_busbw(GB/s)": float(columns[10+i]),
                "ip_wrong":np.int64(columns[11+i]) if columns[11+i].isdigit() or (columns[11+i][1:].isdigit() if columns[11+i].startswith('-') else False) else 0,
            }
        data.append(entry)          
    return data

def concat_dataframes_with_key(df_map: dict[str, pd.DataFrame], column_name: str) -> pd.DataFrame:
    """
    Adds the dictionary key as a new column to each DataFrame,
    and returns a single concatenated DataFrame.

    Args:
        df_map (dict[str, pd.DataFrame]): Dictionary of DataFrames.
        column_name (str): Name of the new column to add.

    Returns:
        pd.DataFrame: Concatenated DataFrame.
    """
    df_list = []
    for key, df in df_map.items():
        df_copy = df.copy()
        df_copy[column_name] = key
        df_list.append(df_copy)
    return pd.concat(df_list, ignore_index=True)

def read_folder_to_DFs(folder_path):
    rvList = {}
    for filename in os.listdir(folder_path):
        filepath = os.path.join(folder_path, filename)
        if os.path.isfile(filepath) and (filename.endswith(".log") or filename.endswith(".txt")):
            data = parse_rccl_tests_output(read_file_as_string(filepath))
            if data:  # Only add sheets if there is data
                df = pd.DataFrame(data)
                rvList[filename] = df
    return rvList


def scatter_df_to_excel_tasks_by_columns(
    df: pd.DataFrame,
    dtype_col: str,
    collective_col: str
) -> list[dict]:
    """
    Splits a DataFrame into smaller DataFrames grouped by data type and collective.
    Returns a list of dicts with filename, sheetname, and the corresponding DataFrame.
    
    Each output dict looks like:
    {
        "filename": f"{input_type}.xlsx",
        "sheetname": collective,
        "df": <DataFrame>
    }
    """
    results = []
    grouped = df.groupby([dtype_col, collective_col])

    for (data_type, collective), group_df in grouped:
        result = {
            "filename": f"{data_type}.xlsx",
            "sheetname": str(collective),
            "df": group_df.reset_index(drop=True)
        }
        results.append(result)

    return results


def power_of_two_to_str(power: int) -> str:
    size_in_bytes = power
    units = ["B", "KB", "MB", "GB", "TB", "PB"]
    for unit in units:
        if size_in_bytes < 1024:
            return f"{int(size_in_bytes)}{unit}"
        size_in_bytes /= 1024
    return f"{int(size_in_bytes)}EB"

def add_human_readable_size_column(df: pd.DataFrame, col_name: str, new_col_name: str = "Size") -> pd.DataFrame:
    df = df.copy()
    readable_sizes = df[col_name].apply(power_of_two_to_str)
    df.insert(0, new_col_name, readable_sizes)
    return df

def fill_merged_block(start_row, end_row, start_col, end_col, text, fill_style, font,border,ws):
    # Auto-unmerge overlapping blocks
    new_range = set(
        (r, c)
        for r in range(start_row, end_row + 1)
        for c in range(start_col, end_col + 1)
    )

    for rng in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = rng.bounds
        existing_range = set(
            (r, c)
            for r in range(min_row, max_row + 1)
            for c in range(min_col, max_col + 1)
        )
        if new_range & existing_range:
            ws.unmerge_cells(str(rng))

    # Merge new block
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=end_row, end_column=end_col
    )

    # Set content and styles for the merged range
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.fill = fill_style
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set text only in the top-left cell
    if text:
        ws.cell(row=start_row, column=start_col).value = text
        
def draw_outer_border_only(ws, start_row, end_row, start_col, end_col, border_style="thin", color="000000"):
    side = Side(border_style=border_style, color=color)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            border = Border()

            if r == start_row:
                border = border + Border(top=side)
            if r == end_row:
                border = border + Border(bottom=side)
            if c == start_col:
                border = border + Border(left=side)
            if c == end_col:
                border = border + Border(right=side)

            cell.border = border

def write_custom_excel_sheet(
    filename,
    coll_name,
    df,
    dir,
    box1_text="Box 1 Default", # BKC,ROCM,RCCL verisons
    box2_text="Box 2 Default", # cmd
    box3_text="Box 3",         # collective info
    box4_text="Box 4",         # out of place
    box5_text="Box 5",         # in place 
    TransferBenchBW = "? GB/s",
    header_row_texts=None
):
    # Ensure directory exists
    os.makedirs(dir, exist_ok=True)
    full_path = os.path.join(dir, filename)

    # Load existing workbook or create new
    if os.path.exists(full_path):
        wb = load_workbook(full_path)
        # Remove existing sheet with the same name
        if coll_name[:31] in wb.sheetnames:
            std = wb[coll_name[:31]]
            wb.remove(std)
    else:
        wb = Workbook()
        # Remove default sheet if file is new
        default = wb.active
        wb.remove(default)

    # Create new sheet
    ws = wb.create_sheet(title=coll_name[:31])

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thick_border = Border(
        left=Side(style="thick", color="000000"),
        right=Side(style="thick", color="000000"),
        top=Side(style="thick", color="000000"),
        bottom=Side(style="thick", color="000000")
    )
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    first_row = 2
    fill_merged_block(first_row, first_row+7, 7, 14, box1_text, yellow_fill,Font(bold=True, color="000000"),thick_border,ws)   # G2:N9
    fill_merged_block(first_row+8, first_row+16, 7, 14, box2_text, yellow_fill,Font(bold=True, color="FF0000"),thick_border,ws) # G10:N18

    # Fill white box groups
    fill_merged_block(first_row+17, first_row+19, 1, 6, box3_text, white_fill,Font(bold=True, color="000000"),thick_border,ws)   # A19:F21
    fill_merged_block(first_row+17, first_row+19, 7, 10, box4_text, white_fill,Font(bold=True, color="000000"),thick_border,ws)  # G19:J21
    fill_merged_block(first_row+17, first_row+19, 11, 14, box5_text, white_fill,Font(bold=True, color="000000"),thick_border,ws) # K19:N21
    # Header row cells A22 to N24
    if header_row_texts is None:
        header_row_texts = [f"H{i+1}" for i in range(14)]
    for col_idx in range(14):
        fill_merged_block(first_row+20, first_row+22, col_idx + 1, col_idx + 1, header_row_texts[col_idx], white_fill,Font(bold=True, color="000000"),thin_border,ws)
    draw_outer_border_only(ws,first_row+20, first_row+22, 1, 6,"thick")
    draw_outer_border_only(ws,first_row+20, first_row+22, 7, 10,"thick")
    draw_outer_border_only(ws,first_row+20, first_row+22, 11, 14,"thick")
    # Write DataFrame starting at A25
    num_rows = df.shape[0]
    data_start_row = first_row+23
    data_end_row = data_start_row+num_rows
    draw_outer_border_only(ws,data_start_row, data_end_row, 1, 6,"thick")
    draw_outer_border_only(ws,data_start_row, data_end_row, 7, 10,"thick")
    draw_outer_border_only(ws,data_start_row, data_end_row, 11, 14,"thick")
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=data_start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_merged_block(data_end_row,data_end_row+1,1,6,"TransferBench (XGMI) 1GB",white_fill,Font(bold=True, color="000000"),thick_border,ws)
    fill_merged_block(data_end_row,data_end_row+1,7,14,TransferBenchBW,white_fill,Font(bold=True, color="000000"),thick_border,ws)
    # Save workbook
    wb.save(full_path)
    
if __name__ == "__main__":
    folder_path = input("Enter the folder path: ")
    output_excel = os.path.join(folder_path,"output.xlsx")  # Change if needed
    datasetdict = read_folder_to_DFs(folder_path)
    combinedDF = concat_dataframes_with_key(datasetdict,"coll")
    group_cols = ['size','elements','type','redop','root',"coll"]
    combinedDF = combinedDF.groupby(group_cols,as_index=False).mean().sort_values(by=['coll','type','elements'])
    split_data = scatter_df_to_excel_tasks_by_columns(combinedDF,dtype_col='type',collective_col='coll')
    # writeDFToExcel(combinedDF,"combinedDF.xlsx")
    header_row_texts = ["size\n[H]","size\n[B]","count\n(elements)","type","redop","root","time\n(us)","algbw\n(GB/s)","bus\n(GB/s)","#wrong","time\n(us)","algbw\n(GB/s)","bus\n(GB/s)","#wrong"]
    for elem in split_data:
        data = add_human_readable_size_column(elem['df'].drop(columns=["coll"]),"size","size_hr")
        write_custom_excel_sheet(elem['filename'], elem['sheetname'], data, "Temp", box1_text="BKC,ROCM,HIP,RCCL versions",box2_text="cmd",box3_text=f"1-node {elem['sheetname']}",box4_text="out-of-place\n(mean of 10 consecutive runs)",box5_text="in-place\n(mean of 10 consecutive runs)",TransferBenchBW = "319 GB/s",header_row_texts=header_row_texts)
    
